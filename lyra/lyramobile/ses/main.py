import base64
import io # Ses verilerini bellekte işlemek için eklendi
import time
from datetime import datetime

import torch
import torchaudio
import numpy as np # Ses işleme için eklendi
from transformers import AutoProcessor, AutoModelForSpeechSeq2Seq

import pandas as pd # Excel'i ilk kez oluşturmak için
from openpyxl import load_workbook, Workbook # Excel'e verimli satır eklemek için
from openpyxl.utils.exceptions import InvalidFileException # Excel dosya hatalarını yakalamak için

# VS Code için Colab'a özgü kütüphaneler yerine bunlar kullanılacak:
import sounddevice as sd
import soundfile as sf # Ses dosyalarını okumak/yazmak için
from playsound import playsound # Ses çalmak için
import requests # Ses dosyalarını indirmek için
import os # Dosya yolları için

# --- Yapılandırma Ayarları ---
MODEL_ADI = "emre/whisper-medium-turkish-2"
TETIKLEYICI_KAYIT_SURESI_MS = 7000  # Tetikleyici için kayıt süresi
KOMUT_KAYIT_SURESI_MS = 15000  # Komut için kayıt süresi
HEDEF_ORNEKLEME_ORANI = 16000 # Modelin beklediği örnekleme oranı (Hz)
MIKROFON_ORNEKLEME_ORANI = 44100 # Mikrofonun varsayılan örnekleme oranı (gerekirse ayarlayın)
MIKROFON_KANAL_SAYISI = 1 # Mono kayıt

EXCEL_DOSYASI = "komutlar.xlsx"
EXCEL_SUTUNLARI = ["Zaman", "Komut"]

TETIKLEYICI_KELIME = "hey akıllı priz"

# Geçerli komutlar için anahtar kelimeler
# komut_gecerli_mi fonksiyonu bu listelerdeki ifadelerin birebir geçip geçmediğini kontrol eder.
# Örneğin, "ışıkları aç" komutu için hem "ışıkları aç" eylemi hem de bir hedef (örn: "salon") geçmelidir.
GECERLI_EYLEMLER = ["ışıkları aç", "ışıkları kapat"]
GECERLI_HEDEFLER = ["salon", "mutfak", "banyo", "oturma odası", "çocuk odası"]

# Sesli geri bildirimler için URL'ler
AUDIO_FILES_DIR = "audio_feedback" # İndirilecek ses dosyaları için klasör
SES_DOSYALARI = {
    "tetikleyici": {"url": None, "yerel_ad": "beep.wav"}, # URL'ler 404 veriyor, yerel dosya varsayalım
    "efendim": {"url": None, "yerel_ad": "efendim.wav"}, # Efendim sesi için yerel dosya, URL'den indirilmeyecek
    "basarili": {"url": None, "yerel_ad": "success.ogg"}, # URL'ler 404 veriyor, yerel dosya varsayalım
    "hata": {"url": None, "yerel_ad": "error.ogg"} # URL'ler 404 veriyor, yerel dosya varsayalım
}

# Yerel ses dosyası yolları (indirildikten sonra kullanılacak)
YEREL_SES_YOLLARI = {}

# Geçici kaydedilen ses dosyaları için klasör
TEMP_AUDIO_DIR = "temp_recordings"

USER_AGENT_HEADER = {"User-Agent": "AkilliPrizAsistani/1.0 (Python Requests; +https://github.com/kullanici/proje_linki)"} # Örnek bir User-Agent

# --- Model Yükleme ---
print(f"🔄 '{MODEL_ADI}' modeli yükleniyor...")
try:
    processor = AutoProcessor.from_pretrained(MODEL_ADI)
    model = AutoModelForSpeechSeq2Seq.from_pretrained(MODEL_ADI)
    CIHAZ = "cuda" if torch.cuda.is_available() else "cpu"
    model = model.to(CIHAZ)
    print(f"✅ Model başarıyla yüklendi ve '{CIHAZ}' cihazına taşındı.")
except Exception as e:
    print(f"❌ Model yüklenirken bir hata oluştu: {e}")
    print("Program sonlandırılıyor.")
    exit()

# --- Ses Dosyalarını İndirme Fonksiyonu ---
def ses_dosyalarini_indir_ve_hazirla():
    """Geri bildirim ses dosyalarını indirir ve yollarını hazırlar."""
    if not os.path.exists(TEMP_AUDIO_DIR):
        os.makedirs(TEMP_AUDIO_DIR)
        print(f"'{TEMP_AUDIO_DIR}' klasörü oluşturuldu.")
    if not os.path.exists(AUDIO_FILES_DIR):
        os.makedirs(AUDIO_FILES_DIR)
        print(f"'{AUDIO_FILES_DIR}' klasörü oluşturuldu.")

    for key, dosya_bilgisi in SES_DOSYALARI.items():
        yerel_yol = os.path.join(AUDIO_FILES_DIR, dosya_bilgisi["yerel_ad"])
        YEREL_SES_YOLLARI[key] = yerel_yol
        if dosya_bilgisi["url"] and not os.path.exists(yerel_yol): # Sadece URL varsa indirmeyi dene
            try:
                print(f"'{dosya_bilgisi['url']}' adresinden '{yerel_yol}' dosyası indiriliyor...")
                response = requests.get(dosya_bilgisi["url"], stream=True, headers=USER_AGENT_HEADER)
                response.raise_for_status() # HTTP hatalarını kontrol et
                with open(yerel_yol, 'wb') as f:
                    for chunk in response.iter_content(chunk_size=8192):
                        f.write(chunk)
                print(f"✅ '{yerel_yol}' başarıyla indirildi.")
            except requests.exceptions.RequestException as e:
                print(f"❌ '{dosya_bilgisi['url']}' indirilirken hata: {e}. Bu ses dosyası kullanılamayabilir.")
                YEREL_SES_YOLLARI[key] = None # Hata durumunda None ata
        elif not dosya_bilgisi["url"] and not os.path.exists(yerel_yol):
            print(f"⚠️ Yerel ses dosyası bulunamadı (URL belirtilmemiş): {yerel_yol}. Bu ses çalınamayacak.")
            YEREL_SES_YOLLARI[key] = None

# --- Excel Dosyası Yardımcı Fonksiyonları ---
def excel_dosyasini_hazirla(dosya_yolu, sutunlar):
    """Excel dosyasının var olduğundan ve doğru başlığa sahip olduğundan emin olur."""
    try:
        workbook = load_workbook(dosya_yolu)
        sheet = workbook.active
        dosyadaki_baslik = [cell.value for cell in sheet[1]]
        if not dosyadaki_baslik or dosyadaki_baslik != sutunlar:
            # Başlık yoksa veya yanlışsa, dosyayı yeniden oluştur (pandas ile)
            print(f"'{dosya_yolu}' dosyasındaki başlık hatalı. Yeniden oluşturuluyor...")
            raise FileNotFoundError # Yeniden oluşturmayı tetikle
    except (FileNotFoundError, InvalidFileException):
        print(f"'{dosya_yolu}' bulunamadı veya geçersiz. Yenisi oluşturuluyor...")
        df_gecici = pd.DataFrame(columns=sutunlar)
        df_gecici.to_excel(dosya_yolu, index=False, engine='openpyxl')
        print(f"✅ '{dosya_yolu}' oluşturuldu ve başlık yazıldı.")
    except Exception as e:
        print(f"❌ Excel dosyası hazırlanırken bir hata oluştu: {e}")

def excel_e_satir_ekle(dosya_yolu, satir_verisi):
    """Excel dosyasına yeni bir satır ekler."""
    try:
        workbook = load_workbook(dosya_yolu)
        sheet = workbook.active
        sheet.append(satir_verisi)
        workbook.save(dosya_yolu)
    except FileNotFoundError:
        print(f"❌ HATA: Excel dosyası ('{dosya_yolu}') ekleme sırasında bulunamadı. Lütfen önce başlatın.")
    except Exception as e:
        print(f"❌ HATA: Excel dosyasına ('{dosya_yolu}') yazılırken hata oluştu: {e}")

# --- Ses İşleme ve Metne Dönüştürme Fonksiyonu ---
def ses_kaydi_al_ve_metne_cevir(kayit_suresi_ms, dosya_adi_log_icin="kayit", dinle=False):
    """Yerel mikrofondan ses kaydeder, işler ve metne dönüştürür."""
    kayit_suresi_saniye = kayit_suresi_ms / 1000.0
    print(f"🎤 {kayit_suresi_saniye:.1f} saniye boyunca ses kaydediliyor ({dosya_adi_log_icin})... (Ctrl+C ile iptal edebilirsiniz)")
    try:
        kaydedilmis_ses = sd.rec(int(kayit_suresi_saniye * MIKROFON_ORNEKLEME_ORANI), # NumPy array döner
                               samplerate=MIKROFON_ORNEKLEME_ORANI,
                               channels=MIKROFON_KANAL_SAYISI, dtype='float32') # Whisper float32 bekler
        sd.wait()  # Kaydın bitmesini bekle
    except Exception as e:
        print(f"⚠️ Ses kaydı sırasında hata: {e}")
        print("🎤 Mikrofon erişim izniniz olduğundan veya bir mikrofon bağlı olduğundan emin olun.")
        return ""

    if kaydedilmis_ses is None or kaydedilmis_ses.size == 0:
        print("⚠️ Kaydedilmiş ses verisi boş veya alınamadı.")
        return ""
    if np.max(np.abs(kaydedilmis_ses)) < 0.01: # Çok düşük genlikli ses kontrolü (eşik değeri ayarlanabilir)
        print(f"⚠️ Kaydedilen sesin genliği çok düşük: {np.max(np.abs(kaydedilmis_ses)):.4f}. Muhtemelen sessizlik kaydedildi.")
        return ""

    if dinle:
        try:
            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            gecici_ses_dosyasi = os.path.join(TEMP_AUDIO_DIR, f"{dosya_adi_log_icin}_{timestamp_str}.wav")
            sf.write(gecici_ses_dosyasi, kaydedilmis_ses, MIKROFON_ORNEKLEME_ORANI)
            print(f"🎶 Kaydedilen ses çalınıyor: {gecici_ses_dosyasi}")
            playsound(gecici_ses_dosyasi)
        except Exception as e_play_rec:
            print(f"⚠️ Kaydedilen ses çalınırken/kaydedilirken hata: {e_play_rec}")

    print("🎧 Ses işleniyor ve metne dönüştürülüyor...")
    try:
        # NumPy dizisini PyTorch tensor'üne dönüştür
        # sounddevice (frames, channels) şeklinde döner, torchaudio (channels, frames) bekler
        if kaydedilmis_ses.ndim > 1 and kaydedilmis_ses.shape[1] == MIKROFON_KANAL_SAYISI:
             waveform = torch.from_numpy(kaydedilmis_ses.T).float()
        else: # Beklenmedik bir format veya mono ise
             waveform = torch.from_numpy(kaydedilmis_ses.squeeze()).float()
             if waveform.ndim == 1: # Mono ise kanal boyutu ekle
                 waveform = waveform.unsqueeze(0)

        if waveform.shape[0] != MIKROFON_KANAL_SAYISI: # Transpoz veya squeeze sonrası kontrol
            print(f"⚠️ Beklenmedik waveform şekli: {waveform.shape}")
            return ""

        print(f"DEBUG: Waveform şekli (resample öncesi): {waveform.shape}, dtype: {waveform.dtype}")

        # Yeniden örnekle
        waveform = torchaudio.functional.resample(waveform, MIKROFON_ORNEKLEME_ORANI, HEDEF_ORNEKLEME_ORANI)
        if waveform.ndim > 1 and waveform.shape[0] > 1: # Eğer stereo ise mono yap
            waveform = torch.mean(waveform, dim=0, keepdim=True)
        waveform_numpy = waveform.squeeze().numpy()
        print(f"DEBUG: Waveform_numpy şekli (resample sonrası, model öncesi): {waveform_numpy.shape}, dtype: {waveform_numpy.dtype}")

        # Metne dönüştür
        inputs = processor(waveform_numpy, sampling_rate=HEDEF_ORNEKLEME_ORANI, return_tensors="pt")
        input_features = inputs.input_features.to(CIHAZ)
        print(f"DEBUG: input_features şekli: {input_features.shape}, dtype: {input_features.dtype}")

        with torch.no_grad():
            output_ids = model.generate(input_features)
        
        # output_ids'nin içeriğini ve şeklini logla
        print(f"DEBUG: output_ids type: {type(output_ids)}, content: {output_ids}")
        if hasattr(output_ids, 'shape'):
            print(f"DEBUG: output_ids shape: {output_ids.shape}")
        if hasattr(output_ids, 'nelement'):
            print(f"DEBUG: output_ids num elements: {output_ids.nelement()}")

        # output_ids'nin None olup olmadığını ve en az bir elemanı olup olmadığını kontrol et
        if output_ids is None or not hasattr(output_ids, 'nelement') or output_ids.nelement() == 0 or (hasattr(output_ids, 'shape') and output_ids.shape[0] == 0):
            print("⚠️ Modelden anlamlı bir çıktı alınamadı (output_ids boş, geçersiz şekil veya eleman yok).")
            return ""

        # output_ids[0] genellikle tek bir ses dosyası için beklenen çıktıdır.
        cozumlenmis_metin = processor.batch_decode(output_ids, skip_special_tokens=True)[0].strip().lower()
        return cozumlenmis_metin

    except Exception as e: # Ses işleme veya metne dönüştürme sırasında oluşabilecek hataları yakala
        print(f"⚠️ Ses işleme veya metne dönüştürme sırasında hata: {e}")
        return "" # Hata durumunda boş metin döndür

# --- Komut Doğrulama Fonksiyonu ---
def komut_gecerli_mi(komut_metni, hedefler, eylemler):
    """Verilen komutun geçerli hedeflerden ve eylemlerden birini içerip içermediğini kontrol eder."""
    hedef_bulundu = any(hedef in komut_metni for hedef in hedefler)
    eylem_bulundu = any(eylem in komut_metni for eylem in eylemler)
    return hedef_bulundu and eylem_bulundu

# --- Ana İşlev ---
def ana_islev():
    ses_dosyalarini_indir_ve_hazirla() # Program başında ses dosyalarını indir/kontrol et
    excel_dosyasini_hazirla(EXCEL_DOSYASI, EXCEL_SUTUNLARI)

    # playsound için not: .ogg dosyaları Windows'ta ek kodek (örn: ffmpeg PATH'de) gerektirebilir.
    # .wav dosyaları genellikle daha sorunsuz çalışır.
    # Gerekirse SES_DOSYALARI'ndaki .ogg linklerini .wav alternatifleriyle değiştirebilirsiniz.
    print("\n--- Ses Çalma Testi (isteğe bağlı) ---")
    if YEREL_SES_YOLLARI.get("tetikleyici"):
        # Test sesini sadece bir kez çalmak için ana döngü dışına aldık,
        # ancak "tetikleyici" sesi zaten tetikleyici algılandığında çalınıyor.
        # Bu test kısmını kaldırabilir veya farklı bir sesle test edebilirsiniz.
        pass
    print(f"🟢 Sistem hazır. '{TETIKLEYICI_KELIME}' demeniz yeterli...")

    while True:
        print(f"\n🔁 Dinleniyor... (Tetikleyici '{TETIKLEYICI_KELIME}' bekleniyor)")
        algilanan_metin = ses_kaydi_al_ve_metne_cevir(
            TETIKLEYICI_KAYIT_SURESI_MS,
            "tetikleyici_kaydi",
            dinle=True)

        if not algilanan_metin: # Kayıt veya çözümleme başarısız olduysa
            print("❌ Tetikleyici algılanamadı (kayıt/çözümleme hatası).")
            time.sleep(1) # Sürekli hata durumunda döngüyü yavaşlat
            continue
            
        # Kaydedilen sesten algılanan metni yazdır (tetikleyici için)
        print(f"🔍 Algılanan metin (tetikleyici için): \"{algilanan_metin}\"")

        if TETIKLEYICI_KELIME in algilanan_metin:
            print(f"🔔 Tetikleyici ('{TETIKLEYICI_KELIME}') algılandı.")
            # Tetikleyici algılandığında "efendim" sesi çal
            if YEREL_SES_YOLLARI.get("efendim"):
                try:
                    playsound(YEREL_SES_YOLLARI["efendim"])
                except Exception as e_play:
                    print(f"⚠️ 'Efendim' sesi çalınırken hata: {e_play}")

            print("🗣️  Lütfen komutunuzu verin...")
            komut = ses_kaydi_al_ve_metne_cevir(
                KOMUT_KAYIT_SURESI_MS,
                "komut_kaydi",
                dinle=True)

            if not komut: # Komut kaydı/çözümlemesi başarısız olduysa
                print("❌ Komut algılanamadı (kayıt/çözümleme hatası).")
                if YEREL_SES_YOLLARI.get("hata"):
                    try:
                        playsound(YEREL_SES_YOLLARI["hata"])
                    except Exception as e_play:
                        print(f"⚠️ Hata sesi çalınırken hata: {e_play}")
                continue

            # Kaydedilen sesten algılanan metni yazdır (komut için)
            print(f"🔍 Algılanan komut metni: \"{komut}\"")

            if komut_gecerli_mi(komut, GECERLI_HEDEFLER, GECERLI_EYLEMLER):
                zaman_damgasi = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                excel_e_satir_ekle(EXCEL_DOSYASI, [zaman_damgasi, komut])
                print(f"✅ Komut işlendi ve kaydedildi: \"{komut}\"")
                if YEREL_SES_YOLLARI.get("basarili"):
                    try:
                        playsound(YEREL_SES_YOLLARI["basarili"])
                    except Exception as e_play:
                        print(f"⚠️ Başarı sesi çalınırken hata: {e_play}")
            else:
                print(f"❌ Hatalı veya eksik komut: \"{komut}\". Lütfen tekrar deneyin.")
                if YEREL_SES_YOLLARI.get("hata"):
                    try:
                        playsound(YEREL_SES_YOLLARI["hata"])
                    except Exception as e_play:
                        print(f"⚠️ Hata sesi çalınırken hata: {e_play}")
        else:
            # Tetikleyici kelime bulunamadıysa, kullanıcıya bilgi vermeden döngüye devam et
            # print(f"❌ Tetikleyici ('{TETIKLEYICI_KELIME}') algılanmadı.")
            pass

if __name__ == "__main__":
    try:
        ana_islev()
    except KeyboardInterrupt:
        print("\n🛑 Program kullanıcı tarafından sonlandırıldı.")
    except Exception as e:
        print(f"\n💥 Beklenmedik bir hata oluştu: {e}")
